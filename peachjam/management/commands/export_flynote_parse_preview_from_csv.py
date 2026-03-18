import csv
from collections import Counter

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

from peachjam.analysis.flynotes import FlynoteParser


class Command(BaseCommand):
    help = "Export parser-produced flynote paths from a CSV file to an XLSX workbook."
    MAX_LEVEL_COLUMNS = 6

    def add_arguments(self, parser):
        parser.add_argument(
            "--input",
            type=str,
            required=True,
            help="Path to the CSV file to read.",
        )
        parser.add_argument(
            "--output",
            type=str,
            required=True,
            help="Path to the XLSX file to create.",
        )
        parser.add_argument(
            "--flynote-column",
            type=str,
            default="Flynote",
            help="CSV column containing raw flynote text.",
        )
        parser.add_argument(
            "--uri-column",
            type=str,
            default="Document - Coredocument Ptr → Expression Frbr Uri",
            help="CSV column containing the document FRBR URI.",
        )
        parser.add_argument(
            "--id-column",
            type=str,
            default="Document - Coredocument Ptr → ID",
            help="CSV column containing the document ID.",
        )

    def handle(self, *args, **options):
        input_path = options["input"]
        output = options["output"]
        if not input_path.lower().endswith(".csv"):
            raise CommandError("--input must end with .csv")
        if not output.lower().endswith(".xlsx"):
            raise CommandError("--output must end with .xlsx")

        parser = FlynoteParser()
        workbook = Workbook()
        paths_sheet = workbook.active
        paths_sheet.title = "paths"
        summary_sheet = workbook.create_sheet("summary")

        with open(input_path, newline="", encoding="utf-8-sig") as infile:
            reader = csv.DictReader(infile)
            if reader.fieldnames is None:
                raise CommandError("CSV file is missing a header row.")

            self.ensure_column_exists(reader.fieldnames, options["flynote_column"])
            self.ensure_column_exists(reader.fieldnames, options["uri_column"])
            self.ensure_column_exists(reader.fieldnames, options["id_column"])

            path_rows = []
            depth_counter = Counter()
            top_level_counter = Counter()
            document_count = 0
            parsed_document_count = 0
            max_depth = 0

            for row_number, row in enumerate(reader, start=2):
                document_count += 1
                raw_flynote = row.get(options["flynote_column"], "") or ""
                normalised_flynote = parser.normalise_multiline_text(raw_flynote)
                paths = parser.parse(raw_flynote)

                base_row = [
                    row_number,
                    row.get(options["id_column"], ""),
                    row.get(options["uri_column"], ""),
                    raw_flynote,
                    normalised_flynote,
                ]

                if not paths:
                    path_rows.append(
                        base_row + ["", "", *[""] * self.MAX_LEVEL_COLUMNS, 0]
                    )
                    continue

                parsed_document_count += 1
                for path in paths:
                    depth = len(path)
                    max_depth = max(max_depth, depth)
                    depth_counter[depth] += 1
                    top_level_counter[path[0]] += 1
                    padded_path = path[: self.MAX_LEVEL_COLUMNS] + [""] * max(
                        0, self.MAX_LEVEL_COLUMNS - len(path)
                    )
                    path_rows.append(base_row + [path[0], *padded_path, depth])

        headers = [
            "csv_row",
            "document_id",
            "frbr_uri",
            "raw_flynote",
            "normalised_flynote",
            "top_level_flynote",
            *[f"flynote_{index}" for index in range(1, self.MAX_LEVEL_COLUMNS + 1)],
            "flynote_depth",
        ]
        paths_sheet.append(headers)
        for row in path_rows:
            paths_sheet.append([self.clean_cell_value(value) for value in row])

        summary_sheet.append(["metric", "value"])
        summary_sheet.append(["documents_processed", document_count])
        summary_sheet.append(["documents_with_paths", parsed_document_count])
        summary_sheet.append(["max_depth", max_depth])
        summary_sheet.append(["top_level_flynote_count", len(top_level_counter)])
        summary_sheet.append([])
        summary_sheet.append(["depth", "path_count"])
        for depth in sorted(depth_counter):
            summary_sheet.append([depth, depth_counter[depth]])
        summary_sheet.append([])
        summary_sheet.append(["top_level_flynote", "path_count"])
        for topic, count in top_level_counter.most_common():
            summary_sheet.append([self.clean_cell_value(topic), count])
        self.style_workbook(paths_sheet, summary_sheet)
        workbook.save(output)
        self.stdout.write(self.style.SUCCESS(f"Wrote flynote preview to {output}"))

    @staticmethod
    def ensure_column_exists(fieldnames, column_name):
        if column_name not in fieldnames:
            raise CommandError(f"CSV column not found: {column_name}")

    @staticmethod
    def clean_cell_value(value):
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    @classmethod
    def style_workbook(cls, paths_sheet, summary_sheet):
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        summary_fill = PatternFill("solid", fgColor="EAF4E2")
        warning_fill = PatternFill("solid", fgColor="FCE8E6")
        bold_font = Font(bold=True)

        for cell in paths_sheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
        paths_sheet.freeze_panes = "A2"
        paths_sheet.auto_filter.ref = paths_sheet.dimensions

        for cell in summary_sheet[1]:
            cell.fill = summary_fill
            cell.font = bold_font

        for row in summary_sheet.iter_rows():
            if row and row[0].value in {
                "depth",
                "top_level_flynote",
            }:
                for cell in row:
                    cell.fill = header_fill
                    cell.font = bold_font
            if row and row[0].value == "top_level_flynote_count":
                for cell in row:
                    cell.fill = warning_fill
                    cell.font = bold_font
