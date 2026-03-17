from collections import Counter

from django.core.management.base import BaseCommand, CommandError
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

from peachjam.analysis.flynotes import FlynoteParser
from peachjam.models import Judgment


class Command(BaseCommand):
    help = "Export parser-produced flynote paths to an XLSX file without updating database records."
    MAX_LEVEL_COLUMNS = 5

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            required=True,
            help="Path to the XLSX file to create.",
        )
        parser.add_argument(
            "--limit",
            type=int,
            default=0,
            help="Process at most N judgments (0 = all).",
        )
        parser.add_argument(
            "--judgment-id",
            type=int,
            default=None,
            help="Export a single judgment by primary key.",
        )
        parser.add_argument(
            "--start-id",
            type=int,
            default=None,
            help="Start processing from this judgment PK downwards.",
        )

    def handle(self, *args, **options):
        output = options["output"]
        if not output.lower().endswith(".xlsx"):
            raise CommandError("--output must end with .xlsx")

        parser = FlynoteParser()
        judgments = self.get_queryset(options)

        workbook = Workbook()
        paths_sheet = workbook.active
        paths_sheet.title = "paths"
        summary_sheet = workbook.create_sheet("summary")

        path_rows = []
        depth_counter = Counter()
        top_level_counter = Counter()
        classified_top_level_counter = Counter()
        judgment_count = 0
        parsed_judgment_count = 0
        max_depth = 0

        for judgment in judgments.iterator():
            judgment_count += 1
            raw_flynote = judgment.flynote or ""
            normalised_flynote = parser.normalise_multiline_text(raw_flynote)
            paths = parser.parse(raw_flynote)

            if not paths:
                path_rows.append(
                    [
                        judgment.pk,
                        judgment.title,
                        judgment.expression_frbr_uri,
                        judgment.get_absolute_url(),
                        raw_flynote,
                        normalised_flynote,
                        "",
                        "",
                        0,
                    ]
                )
                continue

            parsed_judgment_count += 1
            for index, path in enumerate(paths, start=1):
                depth = len(path)
                max_depth = max(max_depth, depth)
                depth_counter[depth] += 1
                top_level_counter[path[0]] += 1
                classified_top_level = parser.classify_top_level_root(path[0])
                classified_top_level_counter[classified_top_level] += 1
                path_rows.append(
                    [
                        judgment.pk,
                        judgment.title,
                        judgment.expression_frbr_uri,
                        judgment.get_absolute_url(),
                        raw_flynote,
                        normalised_flynote,
                        path[0],
                        classified_top_level,
                        *path[: self.MAX_LEVEL_COLUMNS],
                        depth,
                    ]
                )

        headers = [
            "document_id",
            "title",
            "frbr_uri",
            "url",
            "raw_flynote",
            "normalised_flynote",
            "top_level_flynote",
            "classified_top_level_flynote",
            *[f"flynote_{index}" for index in range(1, self.MAX_LEVEL_COLUMNS + 1)],
            "flynote_depth",
        ]
        paths_sheet.append(headers)

        for row in path_rows:
            padded_row = [
                self.clean_cell_value(value)
                for value in (
                    row[:-1] + [""] * max(0, len(headers) - len(row)) + [row[-1]]
                )
            ]
            paths_sheet.append(padded_row)

        summary_sheet.append(["metric", "value"])
        summary_sheet.append(["judgments_processed", judgment_count])
        summary_sheet.append(["judgments_with_paths", parsed_judgment_count])
        summary_sheet.append(["max_depth", max_depth])
        summary_sheet.append(
            ["top_level_flynote_count", len(classified_top_level_counter)]
        )
        summary_sheet.append([])
        summary_sheet.append(["depth", "path_count"])
        for depth in sorted(depth_counter):
            summary_sheet.append([depth, depth_counter[depth]])
        summary_sheet.append([])
        summary_sheet.append(["top_level_flynote", "path_count"])
        for topic, count in top_level_counter.most_common():
            summary_sheet.append([self.clean_cell_value(topic), count])
        summary_sheet.append([])
        summary_sheet.append(["classified_top_level_flynote", "path_count"])
        for topic, count in classified_top_level_counter.most_common():
            summary_sheet.append([self.clean_cell_value(topic), count])

        self.style_workbook(paths_sheet, summary_sheet)

        workbook.save(output)
        self.stdout.write(self.style.SUCCESS(f"Wrote flynote preview to {output}"))

    @staticmethod
    def clean_cell_value(value):
        if isinstance(value, str):
            return ILLEGAL_CHARACTERS_RE.sub("", value)
        return value

    @classmethod
    def style_workbook(cls, paths_sheet, summary_sheet):
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        summary_fill = PatternFill("solid", fgColor="EAF4E2")
        warning_fill = PatternFill("solid", fgColor="FCE8E6")
        bold_font = Font(bold=True)

        for cell in paths_sheet[1]:
            cell.fill = header_fill
            cell.font = bold_font
        paths_sheet.freeze_panes = "A2"
        paths_sheet.auto_filter.ref = paths_sheet.dimensions

        for cell in summary_sheet[1]:
            cell.fill = summary_fill
            cell.font = bold_font

        for row in summary_sheet.iter_rows():
            if row and row[0].value == "top_level_flynote":
                for cell in row:
                    cell.fill = header_fill
                    cell.font = bold_font
            if row and row[0].value == "classified_top_level_flynote":
                for cell in row:
                    cell.fill = header_fill
                    cell.font = bold_font
            if row and row[0].value == "top_level_flynote_count":
                for cell in row:
                    cell.fill = warning_fill
                    cell.font = bold_font

    def get_queryset(self, options):
        if options["judgment_id"]:
            judgment = Judgment.objects.filter(pk=options["judgment_id"]).first()
            if not judgment:
                raise CommandError(
                    f"Judgment with pk={options['judgment_id']} not found."
                )
            return Judgment.objects.filter(pk=judgment.pk)

        queryset = (
            Judgment.objects.exclude(flynote__isnull=True)
            .exclude(flynote="")
            .order_by("-pk")
        )

        if options["start_id"]:
            queryset = queryset.filter(pk__lte=options["start_id"])

        if options["limit"]:
            queryset = queryset[: options["limit"]]

        return queryset
